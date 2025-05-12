from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from .models import CadastroAlunos, RegAtrasos, Presenca
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import locale
import calendar

#windows
locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')

# direciona para a home
@login_required
def home(request):
    return render(request, "app_gestao/home.html")

# realiza o cadastro das turmas, atravez do arquivo .xlsx
@login_required
def cadastro(request):
    if request.method == 'POST' and request.FILES['arquivo_xlsx']:
        arquivo_xlsx = request.FILES['arquivo_xlsx']
        try:
            df = pd.read_excel(arquivo_xlsx)  # Lê o arquivo Excel em um DataFrame

            # Certifica-se de que as colunas do Excel correspondam aos nomes das colunas do banco de dados
            df.rename(columns={
                'R.A.': 'ra',
                'Nome do estudante': 'nome_estudante',
                'Série/turma': 'serie_turma',
                'Endereço': 'endereco',
                'Responsável 1': 'responsavel1',
                'Responsável 2': 'responsavel2',
                'Contato(s)': 'contato'
            }, inplace=True)
            for index, row in df.iterrows():
                CadastroAlunos.objects.update_or_create(
                    ra=row['ra'],
                    defaults={
                        'nome_estudante': row['nome_estudante'],
                        'serie_turma': row['serie_turma'],
                        'endereco': row['endereco'],
                        'responsavel1': row['responsavel1'],
                        'responsavel2': row['responsavel2'],
                        'contato': row['contato'],
                    }
                )
            messages.success(request, 'Dados importados com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao importar dados: {e}')   
        return redirect('cadastro')
    return render(request, 'app_gestao/cadastrar_alunos.html')

# realiza o registro da presença dos alunos
@login_required
def registrar_presenca(request):
    data_selecionada = request.POST.get("data") or request.GET.get("data")
    turma_selecionada = request.POST.get("turma") or request.GET.get("turma")
    alunos = []

    if turma_selecionada:
        alunos = CadastroAlunos.objects.filter(serie_turma=turma_selecionada)

    if request.method == "POST" and "registrar" in request.POST:
        for aluno in alunos:
            status = request.POST.get(f"presenca_{aluno.ra}", "presente")
            presente = status == "presente"
            try:
                Presenca.objects.update_or_create(
                    aluno=aluno,
                    data=data_selecionada,
                    defaults={"presente": presente}
                )
            except IntegrityError:
                messages.error(request, f"Erro ao registrar presença para {aluno.nome_estudante}")

        messages.success(request, "Presenças registradas com sucesso!")
        return redirect("registrar_presenca")

    turmas = CadastroAlunos.objects.values_list("serie_turma", flat=True).distinct()

    return render(request, "app_gestao/registrar_presenca.html", {
        "data_selecionada": data_selecionada,
        "turma_selecionada": turma_selecionada,
        "alunos": alunos,
        "turmas": turmas,
    })

#Limpa os dados dos bancos de dados
@login_required
def limpar_banco(request):
    if request.method == 'POST':
        CadastroAlunos.objects.all().delete()
        RegAtrasos.objects.all().delete()
        Presenca.objects.all().delete()
        messages.success(request, 'Todos os dados foram apagados com sucesso.')
    return redirect('cadastro')  # substitua com o nome da sua view

#seleciona o aluno para o registro do atraso
@login_required
def registrar_atraso(request):
    turmas = CadastroAlunos.objects.values_list('serie_turma', flat=True).distinct()
    alunos = []
    turma_selecionada = None

    if request.method == 'POST':
        turma_selecionada = request.POST.get('turma')
        if turma_selecionada:
            alunos = CadastroAlunos.objects.filter(serie_turma=turma_selecionada)

    return render(request, 'app_gestao/registrar_atraso.html', {
        'turmas': turmas,
        'alunos': alunos,
        'turma_selecionada': turma_selecionada,
    })

#efetua o registro do atraso do aluno selecionado no banco de dados
@login_required
def registrar_atraso_aluno(request, ra):
    aluno = get_object_or_404(CadastroAlunos, ra=ra)

    if request.method == 'POST' and 'salvar' in request.POST:
        data_atraso = request.POST.get('data_atraso')
        horario_chegada = request.POST.get('horario_chegada')
        justificativa = request.POST.get('justificativa')

        RegAtrasos.objects.create(
            ra=aluno,
            data_atraso=data_atraso,
            horario_chegada=horario_chegada,
            justificativa=justificativa,
        )
        messages.success(request, 'Atraso registrado com sucesso!')
        return redirect('registrar_atraso')

    return render(request, 'app_gestao/registrar_atraso_aluno.html', {'aluno': aluno})

# apresenta o relatório com o atrasos dos alunos de acordo com a turma
@login_required
def relatorio(request):
    mes = request.GET.get('mes')
    turma = request.GET.get('turma')
    exportar = request.GET.get('exportar')
    relatorio = []
    nome_mes = None

    meses = [(i, calendar.month_name[i]) for i in range(1, 13)]
    turmas = CadastroAlunos.objects.values_list('serie_turma', flat=True).distinct()

    if mes and turma:
        mes = int(mes)
        alunos = CadastroAlunos.objects.filter(serie_turma=turma)
        nome_mes = calendar.month_name[int(mes)].capitalize()

        for aluno in alunos:
            presencas = Presenca.objects.filter(
                aluno_id=aluno.ra,
                data__month=mes,
                presente=1
            ).count()

            atrasos = RegAtrasos.objects.filter(
                ra=aluno.ra,
                data_atraso__month=mes
            ).count()

            percentual = round((atrasos / presencas) * 100, 2) if presencas > 0 else 0
            cor = ''
            if percentual > 50:
                cor = 'danger'
            elif percentual > 0:
                cor = 'warning'

            relatorio.append({
                'aluno': aluno,
                'turma': turma,
                'mes': nome_mes,
                'presencas': presencas,
                'atrasos': atrasos,
                'percentual': percentual,
                'cor': cor,
            })

        if exportar == "1":
            # Gerar planilha Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Relatório"

            # Cabeçalhos
            ws.append(["Aluno", "Turma", "Mes", "Presenças", "Atrasos", "% Atraso"])

            # Total de colunas do relatório
            num_colunas = 6

            for item in relatorio:
                row = [
                    item['aluno'].nome_estudante,
                    item['turma'],
                    item['mes'],
                    item['presencas'],
                    item['atrasos'],
                    item['percentual'],
                ]
                ws.append(row)

                # Cor da linha inteira conforme a cor do item
                if item['cor'] == 'danger':
                    fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                elif item['cor'] == 'warning':
                    fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                else:
                    fill = None

                if fill:
                    for col in range(1, num_colunas + 1):
                        ws.cell(row=ws.max_row, column=col).fill = fill

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=relatorio.xlsx'
            wb.save(response)
            return response

    context = {
        'mes': mes,
        'turma': turma,
        'meses': meses,
        'turmas': turmas,
        'relatorio': relatorio,
        'nome_mes': nome_mes
    }
    return render(request, 'app_gestao/relatorio.html', context)

#relatório de detalhes de atrasos do aluno
@login_required
def detalhes_atrasos(request, ra):
    mes = request.GET.get('mes')
    turma = request.GET.get('turma')
    aluno = get_object_or_404(CadastroAlunos, ra=ra)
    nome_mes = None

    #meses = [(i, calendar.month_name[i]) for i in range(1, 13)]
    #turmas = CadastroAlunos.objects.values_list('serie_turma', flat=True).distinct()

    if mes and turma:
        mes = int(mes)
        #alunos = CadastroAlunos.objects.filter(serie_turma=turma)
        nome_mes = calendar.month_name[int(mes)].capitalize()

    atrasos = RegAtrasos.objects.filter(
        ra=ra,
        data_atraso__month=mes
    ).order_by('data_atraso')

    context = {
        'aluno': aluno,
        'turma': turma,
        'mes': nome_mes,
        'atrasos': atrasos
    }
    return render(request, 'app_gestao/detalhes_atrasos.html', context)
